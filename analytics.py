import os
from datetime import datetime
from pathlib import Path
from threading import Lock
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

LOCK = Lock()

from pathlib import Path

# Onde o drive 'dtic' foi montado no Linux
# MOUNT_POINT = Path("/mnt/z")
MOUNT_POINT = Path("Z:/")
# A pasta específica dentro do compartilhamento
BASE_DIR = MOUNT_POINT / "PDF Otimizador"

# Estrutura de dados
DATA_DIR = BASE_DIR
USO_XLSX = DATA_DIR / "acessos" / "uso.xlsx"
FEEDBACK_XLSX = DATA_DIR / "feedbacks" / "feedback.xlsx"

# Garantir que as pastas existam antes de salvar os arquivos
try:
    USO_XLSX.parent.mkdir(parents=True, exist_ok=True)
    FEEDBACK_XLSX.parent.mkdir(parents=True, exist_ok=True)
    print(f"Sucesso! Caminho configurado: {BASE_DIR}")
except PermissionError:
    print("Erro de permissão: Verifique se o 'Mr robot' tem acesso de escrita no Windows.")
    

HEADERS = ["Data", "Hora", "Usuário", "Ação", "Módulo", "Tempo", "Estrelas", "Descrição", "IP", "Observação"]

print(f"Diretório de uso criado em: {USO_XLSX.parent}")

def _ensure_parent_dir(file_path: str):
    parent = os.path.dirname(os.path.abspath(file_path))
    os.makedirs(parent, exist_ok=True)


def _agora_data_hora():
    now = datetime.now()
    return now.strftime("%d/%m/%Y"), now.strftime("%H:%M:%S")


def _setup_sheet(ws):
    ws.append(HEADERS)

    bold = Font(bold=True)
    center = Alignment(vertical="center")

    for cell in ws[1]:
        cell.font = bold
        cell.alignment = center

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
    ws.row_dimensions[1].height = 18

    widths = [12, 10, 18, 14, 16, 10, 10, 45, 16, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _parse_float(value):
    try:
        if value is None:
            return None
        return float(str(value).strip().replace(",", "."))
    except Exception:
        return None


def _parse_int(value):
    try:
        return int(str(value).strip())
    except Exception:
        return None


def _parse_kv(text: str):
    out = {}
    for part in (text or "").split(";"):
        if "=" not in part:
            continue
        key, value = part.split("=", 1)
        out[key.strip()] = value.strip()
    return out


def _looks_like_ip(value):
    if not isinstance(value, str):
        return False
    parts = value.strip().split(".")
    if len(parts) != 4:
        return False
    for part in parts:
        if not part.isdigit():
            return False
        if not 0 <= int(part) <= 255:
            return False
    return True


def _append_xlsx(path: str, values: list):
    _ensure_parent_dir(path)

    with LOCK:
        if not os.path.exists(path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Dados"
            _setup_sheet(ws)
            ws.append(values)
            wb.save(path)
            return

        wb = load_workbook(path)
        ws = wb.active

        if ws.max_row < 1 or (ws.max_row == 1 and ws["A1"].value != "Data"):
            ws.delete_rows(1, ws.max_row)
            _setup_sheet(ws)

        ws.append(values)
        wb.save(path)


# =========================
# USO (uso.xlsx)
# =========================
def log_uso(
    acao: str,
    modulo: str,
    ip: str,
    usuario: str = "",
    descricao: str = "",
    observacao: str = "",
    tempo = "",
    estrelas = "",
):
    data, hora = _agora_data_hora()
    values = [
        data,
        hora,
        usuario or "",
        (acao or "").strip(),
        (modulo or "").strip(),
        tempo if tempo is not None else "",
        estrelas if estrelas is not None else "",
        (descricao or "").strip(),
        ip or "",
        (observacao or "").strip(),
    ]
    _append_xlsx(USO_XLSX, values)


# =========================
# FEEDBACK (feedback.xlsx)
# =========================
def log_feedback(
    estrelas: int,
    descricao: str,
    ip: str,
    usuario: str = "",
    modulo: str = "",
    tempo: int = "",
    observacao: str = "",
):
    data, hora = _agora_data_hora()

    try:
        estrelas = int(estrelas)
    except Exception:
        estrelas = 0

    values = [
        data,
        hora,
        usuario or "",
        "feedback",
        (modulo or "").strip(),
        tempo if tempo is not None else "",
        estrelas,
        (descricao or "").strip(),
        ip or "",
        (observacao or "").strip(),
    ]
    _append_xlsx(FEEDBACK_XLSX, values)


def _read_all_xlsx(path: str):
    if not os.path.exists(path):
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    headers = [h for h in rows[0]]
    out = []
    for r in rows[1:]:
        if not r:
            continue
        item = {}
        for i, h in enumerate(headers):
            item[str(h)] = r[i] if i < len(r) else ""
        out.append(item)
    return out


def read_tail_uso(limit: int = 200):
    rows = _read_all_xlsx(USO_XLSX)
    return _normalize_uso_rows(rows[-limit:])


def read_tail_feedback(limit: int = 200):
    rows = _read_all_xlsx(FEEDBACK_XLSX)
    return _normalize_feedback_rows(rows[-limit:])


def _normalize_uso_rows(rows):
    out = []
    for r in rows:
        data = r.get("Data", "")
        hora = r.get("Hora", "")
        acao = r.get("Ação", "")
        modulo = r.get("Módulo", "")
        tempo = r.get("Tempo", "")
        descricao = r.get("Descrição", "")
        ip = r.get("IP", "")
        observacao = r.get("Observação", "")

        if not observacao and isinstance(ip, str) and "=" in ip:
            observacao = ip
            ip = ""

        if (not ip or not _looks_like_ip(ip)) and _looks_like_ip(descricao):
            ip = descricao
            descricao = ""

        kv = _parse_kv(observacao)
        if not tempo:
            tempo = kv.get("secs", "")
        in_mb = _parse_float(kv.get("in_mb"))
        out_mb = _parse_float(kv.get("out_mb"))
        if in_mb is None:
            in_mb = ""
        if out_mb is None:
            out_mb = ""
        out.append(
            {
                "ts": f"{data} {hora}".strip(),
                "event": acao,
                "module": modulo,
                "tempo": tempo,
                "endpoint": kv.get("endpoint", ""),
                "task_id": kv.get("id_tarefa", ""),
                "in_mb": in_mb,
                "out_mb": out_mb,
                #"ip": ip,
                "extra": descricao or observacao,
                "observacao": observacao,
            }
        )
    return out


def _normalize_feedback_rows(rows):
    out = []
    for r in rows:
        data = r.get("Data", "")
        hora = r.get("Hora", "")
        modulo = r.get("Módulo", "")
        tempo = r.get("Tempo", "")
        estrelas = r.get("Estrelas", "")
        descricao = r.get("Descrição", "")
        ip = r.get("IP", "")

        stars_val = _parse_int(estrelas)
        if not stars_val or stars_val not in range(1, 6):
            stars_val = _parse_int(tempo) or 0

        message = descricao
        if _looks_like_ip(descricao) and isinstance(estrelas, str) and estrelas:
            message = estrelas
        if not message and isinstance(estrelas, str):
            message = estrelas

        if not _looks_like_ip(ip) and _looks_like_ip(descricao):
            ip = descricao

        out.append(
            {
                "ts": f"{data} {hora}".strip(),
                "stars": stars_val,
                "module": modulo,
                "message": message,
                "ip": ip,
            }
        )
    return out


def compute_metrics():
    uso = _read_all_xlsx(USO_XLSX)
    feedback = _read_all_xlsx(FEEDBACK_XLSX)

    def norm(x):
        return (x or "").strip().lower()

    uploads = sum(1 for r in uso if norm(r.get("Ação")) == "upload")
    downloads = sum(1 for r in uso if norm(r.get("Ação")) == "download")
    inicios = sum(1 for r in uso if norm(r.get("Ação")) in ["início", "inicio"])
    concluidos = sum(1 for r in uso if norm(r.get("Ação")) in ["concluído", "concluido"])
    erros = sum(1 for r in uso if norm(r.get("Ação")) == "erro")
    cancelados = sum(1 for r in uso if norm(r.get("Ação")) == "cancelado")

    estrelas_list = []
    for r in feedback:
        stars_val = _parse_int(r.get("Estrelas"))
        if not stars_val or stars_val not in range(1, 6):
            stars_val = _parse_int(r.get("Tempo")) or 0
        if stars_val:
            estrelas_list.append(stars_val)

    media = round(sum(estrelas_list) / len(estrelas_list), 2) if estrelas_list else 0.0
    aprovacoes = sum(1 for s in estrelas_list if s >= 4)
    taxa_aprov = round((aprovacoes / len(estrelas_list)) * 100, 1) if estrelas_list else 0.0

    upload_sizes = []
    input_sizes = []
    output_sizes = []
    times = []

    for r in uso:
        acao = norm(r.get("Ação"))
        observacao = r.get("Observação") or ""
        if not observacao:
            ip_col = r.get("IP")
            if isinstance(ip_col, str) and "=" in ip_col:
                observacao = ip_col

        kv = _parse_kv(observacao)

        if acao == "upload":
            in_mb = _parse_float(kv.get("in_mb"))
            if in_mb is None:
                in_mb = _parse_float(kv.get("tamanho_mb"))
            if in_mb:
                upload_sizes.append(in_mb)

        if acao in ["concluído", "concluido"]:
            in_mb = _parse_float(kv.get("in_mb"))
            out_mb = _parse_float(kv.get("out_mb"))
            if in_mb:
                input_sizes.append(in_mb)
            if out_mb:
                output_sizes.append(out_mb)

            tempo_val = _parse_float(r.get("Tempo"))
            if tempo_val is None:
                tempo_val = _parse_float(kv.get("secs"))
            if tempo_val:
                times.append(tempo_val)

    total_upload_mb = round(sum(upload_sizes), 2) if upload_sizes else 0.0
    avg_upload_mb = round(total_upload_mb / len(upload_sizes), 2) if upload_sizes else 0.0

    total_input_mb = round(sum(input_sizes), 2) if input_sizes else 0.0
    avg_input_mb = round(total_input_mb / len(input_sizes), 2) if input_sizes else 0.0
    total_output_mb = round(sum(output_sizes), 2) if output_sizes else 0.0
    avg_output_mb = round(total_output_mb / len(output_sizes), 2) if output_sizes else 0.0

    saved_total_mb = round(total_input_mb - total_output_mb, 2) if total_input_mb else 0.0
    reduction_pct = round((1 - (total_output_mb / total_input_mb)) * 100, 1) if total_input_mb else 0.0

    avg_time_sec = round(sum(times) / len(times), 2) if times else 0.0
    max_time_sec = round(max(times), 2) if times else 0.0

    return {
        "uploads": uploads,
        "downloads": downloads,
        "inicios": inicios,
        "concluidos": concluidos,
        "erros": erros,
        "cancelados": cancelados,
        "feedback_count": len(feedback),
        "media_estrelas": media,
        "taxa_aprovacao_pct": taxa_aprov,
        "aprovacoes_4_ou_5": aprovacoes,
        "avg_stars": media,
        "approval_rate_pct": taxa_aprov,
        "total_upload_mb": total_upload_mb,
        "avg_upload_mb": avg_upload_mb,
        "total_input_mb": total_input_mb,
        "avg_input_mb": avg_input_mb,
        "total_output_mb": total_output_mb,
        "avg_output_mb": avg_output_mb,
        "saved_total_mb": saved_total_mb,
        "reduction_pct": reduction_pct,
        "avg_time_sec": avg_time_sec,
        "max_time_sec": max_time_sec,
    }
